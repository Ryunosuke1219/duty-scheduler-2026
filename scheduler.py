#!/usr/bin/env python3
"""
当直シフトスケジューラ v2.0
OR-Tools CP-SAT Solverを使用した最適化アルゴリズム

使い方:
    python scheduler.py <CSV_FILE> <YEAR> <MONTH>
    例: python scheduler.py 2026_1月_shift_schedule.csv 2026 1
"""

import sys
import re
import datetime
from collections import defaultdict
from pathlib import Path

try:
    import pandas as pd
    from ortools.sat.python import cp_model
    import openpyxl
    from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print("必要なパッケージがインストールされていません。")
    print("以下のコマンドを実行してください:")
    print()
    print("  pip install ortools pandas openpyxl")
    print()
    sys.exit(1)


def parse_shift_column(col: str) -> tuple[int, int]:
    """シフト列名から日付とサブシフトを抽出

    Returns:
        (day, sub): day=日付, sub=0(通常当直), 1(-1:日直), 2(-2:当直)
    """
    m = re.match(r'(\d+)', col)
    if not m:
        return None, None
    day = int(m.group(1))
    sub_match = re.search(r'-(\d)', col)
    sub = int(sub_match.group(1)) if sub_match else 0
    return day, sub


def get_weekday(col: str) -> str:
    """シフト列名から曜日を抽出"""
    m = re.search(r'\((\w+)\)', col)
    return m.group(1) if m else ""


class DutyScheduler:
    def __init__(self, csv_path: str, year: int, month: int):
        self.year = year
        self.month = month
        self.csv_path = csv_path

        # データ読み込み
        self.df_raw = pd.read_csv(csv_path, encoding='cp932')
        self.doctors = self.df_raw['Name'].dropna().tolist()
        self.doctors = [d for d in self.doctors if d]  # 空文字除去
        self.group_map = self.df_raw.set_index('Name')['Group'].to_dict()

        # 事前割り当ての検出 (3=当直, 4=OC)
        self.pre_duty = {}  # {shift: {doc: group}}
        self.pre_oc = {}    # {shift: doc}
        data_cols = self.df_raw.drop(columns=['Group', 'Name'])
        for idx, row in self.df_raw.iterrows():
            doc = row['Name']
            if pd.isna(doc) or not doc:
                continue
            group = 'G0' if row['Group'] == 0 else 'G1'
            for col in data_cols.columns:
                val = row[col]
                if val == 3:  # 事前当直割り当て
                    if col not in self.pre_duty:
                        self.pre_duty[col] = {}
                    self.pre_duty[col][doc] = group
                elif val == 4:  # 事前OC割り当て
                    self.pre_oc[col] = doc

        if self.pre_duty or self.pre_oc:
            print(f"事前割り当て検出: 当直{len(self.pre_duty)}シフト, OC{len(self.pre_oc)}シフト")

        # 利用可能性マトリクス (1=可, 0=不可に反転、3/4も可として扱う)
        avail_raw = self.df_raw.drop(columns=['Group', 'Name']).copy()
        # 0→1(可), 1→0(不可), 3→1(可), 4→1(可)
        self.avail_df = avail_raw.apply(lambda x: x.apply(lambda v: 1 if v in [0, 3, 4] else 0))
        self.avail_df.index = self.df_raw['Name']

        # シフト列の解析
        self.shift_cols = [c for c in self.avail_df.columns
                          if parse_shift_column(c)[0] is not None]
        self.shift_cols = sorted(self.shift_cols,
                                 key=lambda c: (parse_shift_column(c)[0],
                                               parse_shift_column(c)[1]))

        # 各シフトの属性
        self.shift_day = {}
        self.shift_sub = {}
        self.shift_date = {}
        self.shift_week = {}

        for col in self.shift_cols:
            day, sub = parse_shift_column(col)
            self.shift_day[col] = day
            self.shift_sub[col] = sub
            try:
                date = datetime.date(year, month, day)
                self.shift_date[col] = date
                self.shift_week[col] = date.isocalendar().week
            except ValueError:
                continue

        # 祝日判定: 土曜以外で -1/-2 シフトがある日
        self.holidays = set()
        for col in self.shift_cols:
            if col not in self.shift_date:
                continue
            date = self.shift_date[col]
            if self.shift_sub[col] > 0 and date.weekday() != 5:  # 土曜以外
                if date.weekday() != 6:  # 日曜以外 = 祝日
                    self.holidays.add(date)

        # グループ分け
        self.g0 = [d for d in self.doctors if self.group_map.get(d) == 0]
        self.g1 = [d for d in self.doctors if self.group_map.get(d) == 1]

        # シフト分類
        self.weekday_shifts = []  # 平日シフト
        self.weekend_shifts = []  # 土日祝シフト
        self.sunday_holiday_nichoku = []  # 日曜・祝日の日直(-1)
        self.saturday_nichoku = []  # 土曜の日直(-1)

        for col in self.shift_cols:
            if col not in self.shift_date:
                continue
            date = self.shift_date[col]
            sub = self.shift_sub[col]

            is_weekend = date.weekday() >= 5  # 土日
            is_holiday = date in self.holidays

            if is_weekend or is_holiday:
                self.weekend_shifts.append(col)
                if sub == 1:  # 日直
                    if date.weekday() == 6 or is_holiday:
                        self.sunday_holiday_nichoku.append(col)
                    elif date.weekday() == 5:
                        self.saturday_nichoku.append(col)
            else:
                self.weekday_shifts.append(col)

        print(f"=== 読み込み完了 ===")
        print(f"医師数: G0={len(self.g0)}人, G1={len(self.g1)}人")
        print(f"シフト数: 平日={len(self.weekday_shifts)}, 土日祝={len(self.weekend_shifts)}")
        print(f"祝日: {sorted([d.day for d in self.holidays])}日")

    def is_available(self, doc: str, shift: str) -> bool:
        """医師がそのシフトに勤務可能か"""
        if doc not in self.avail_df.index:
            return False
        if shift not in self.avail_df.columns:
            return False
        return self.avail_df.loc[doc, shift] == 1

    def solve(self, time_limit: int = 60) -> dict:
        """CP-SATソルバーで最適化"""
        model = cp_model.CpModel()

        # === 変数定義 ===
        # x[doc][shift] = 1 なら doc が shift に当直
        x = {}
        for doc in self.doctors:
            x[doc] = {}
            for shift in self.shift_cols:
                x[doc][shift] = model.NewBoolVar(f'x_{doc}_{shift}')

        # oc[doc][shift] = 1 なら doc が shift のOC
        oc = {}
        for doc in self.g0:  # OCはG0のみ
            oc[doc] = {}
            for shift in self.shift_cols:
                oc[doc][shift] = model.NewBoolVar(f'oc_{doc}_{shift}')

        # === ハード制約 ===

        # 0. 事前割り当ての固定
        for shift, doctors in self.pre_duty.items():
            if shift not in self.shift_cols:
                continue
            for doc, group in doctors.items():
                if doc in self.doctors:
                    model.Add(x[doc][shift] == 1)
                    # 他の医師はこのシフトに入れない（同グループ内）
                    if group == 'G0':
                        for other in self.g0:
                            if other != doc:
                                model.Add(x[other][shift] == 0)
                    else:
                        for other in self.g1:
                            if other != doc:
                                model.Add(x[other][shift] == 0)

        for shift, doc in self.pre_oc.items():
            if shift not in self.shift_cols:
                continue
            if doc in self.g0:
                model.Add(oc[doc][shift] == 1)
                # 他のG0はこのシフトのOCに入れない
                for other in self.g0:
                    if other != doc:
                        model.Add(oc[other][shift] == 0)

        # 1. 不可日は割り当て禁止
        for doc in self.doctors:
            for shift in self.shift_cols:
                if not self.is_available(doc, shift):
                    model.Add(x[doc][shift] == 0)

        # OC も不可日禁止
        for doc in self.g0:
            for shift in self.shift_cols:
                if not self.is_available(doc, shift):
                    model.Add(oc[doc][shift] == 0)

        # 2. 各シフトに必要人数を割り当て
        # 日直(-1シフト): G0とG1の両方から1人ずつ
        # それ以外: G0またはG1から1人（どちらでもよい）
        # ※事前割り当て済みシフトは除外
        nichoku_shifts = [s for s in self.shift_cols if self.shift_sub.get(s) == 1]
        other_shifts = [s for s in self.shift_cols if self.shift_sub.get(s) != 1]

        for shift in nichoku_shifts:
            if shift in self.pre_duty:
                # 事前割り当て済み: 不足分のみ補充
                pre_g0 = [d for d, g in self.pre_duty[shift].items() if g == 'G0']
                pre_g1 = [d for d, g in self.pre_duty[shift].items() if g == 'G1']
                if not pre_g0:
                    model.Add(sum(x[d][shift] for d in self.g0) == 1)
                if not pre_g1:
                    model.Add(sum(x[d][shift] for d in self.g1) == 1)
            else:
                model.Add(sum(x[d][shift] for d in self.g0) == 1)
                model.Add(sum(x[d][shift] for d in self.g1) == 1)

        for shift in other_shifts:
            if shift in self.pre_duty:
                # 事前割り当て済み: すでに人がいるのでスキップ
                continue
            # G0またはG1から合計1人
            model.Add(sum(x[d][shift] for d in self.doctors) == 1)

        # 3. 週1回制限（ハード制約）
        # ※事前割り当て済みシフトは除外（年末年始など特殊ケース対応）
        weeks = set(self.shift_week.values())
        for doc in self.doctors:
            for week in weeks:
                week_shifts = [s for s in self.shift_cols
                              if self.shift_week.get(s) == week]
                # 事前割り当て済みシフトを除外
                free_shifts = [s for s in week_shifts
                              if s not in self.pre_duty or doc not in self.pre_duty.get(s, {})]
                if free_shifts:
                    model.Add(sum(x[doc][s] for s in free_shifts) <= 1)

        # 4. 日曜・祝日の日直は月1回まで
        for doc in self.doctors:
            model.Add(sum(x[doc][s] for s in self.sunday_holiday_nichoku) <= 1)

        # 5. 個別制約: 小波津は日直1回のみ、OC0回
        if '小波津' in self.doctors:
            # 日直(-1シフト)のみ1回、それ以外は0回
            model.Add(sum(x['小波津'][s] for s in nichoku_shifts) <= 1)
            model.Add(sum(x['小波津'][s] for s in other_shifts) == 0)
            # OC禁止
            if '小波津' in self.g0:
                for shift in self.shift_cols:
                    model.Add(oc['小波津'][shift] == 0)

        # 5. G1が当直するシフト(日直以外)にはG0からOC
        # ※事前割り当て済みOCがあるシフトはスキップ
        for shift in other_shifts:
            if shift in self.pre_oc:
                # 事前割り当て済みOCがある場合はスキップ
                continue
            if shift in self.pre_duty:
                # 事前割り当て済み当直がある場合もスキップ（OC含め固定済み）
                continue

            # G1が割り当てられた場合のみOC必要
            g1_vars = [x[d][shift] for d in self.g1]
            oc_vars = [oc[d][shift] for d in self.g0]

            # G1が担当 → OCが必要
            g1_assigned = model.NewBoolVar(f'g1_assigned_{shift}')
            model.Add(sum(g1_vars) >= 1).OnlyEnforceIf(g1_assigned)
            model.Add(sum(g1_vars) == 0).OnlyEnforceIf(g1_assigned.Not())

            # G1が担当ならOC1人必要、そうでなければOC不要
            model.Add(sum(oc_vars) == 1).OnlyEnforceIf(g1_assigned)
            model.Add(sum(oc_vars) == 0).OnlyEnforceIf(g1_assigned.Not())

        # 6. OCは当直者本人と重複不可
        for doc in self.g0:
            for shift in self.shift_cols:
                model.Add(x[doc][shift] + oc[doc][shift] <= 1)

        # === ソフト制約 (目的関数) ===
        penalties = []

        # 平日当直回数カウント
        weekday_count = {}
        for doc in self.doctors:
            weekday_count[doc] = sum(x[doc][s] for s in self.weekday_shifts)

        # 土日祝当直回数カウント
        weekend_count = {}
        for doc in self.doctors:
            weekend_count[doc] = sum(x[doc][s] for s in self.weekend_shifts)

        # OC回数カウント
        oc_count = {}
        for doc in self.g0:
            oc_count[doc] = sum(oc[doc][s] for s in self.shift_cols)

        # ペナルティ1: 平日1回から乖離
        for doc in self.doctors:
            dev_weekday = model.NewIntVar(0, 10, f'dev_weekday_{doc}')
            model.AddAbsEquality(dev_weekday, weekday_count[doc] - 1)
            penalties.append(dev_weekday * 100)  # 高い重み

        # ペナルティ2: 土日祝1回から乖離
        for doc in self.doctors:
            dev_weekend = model.NewIntVar(0, 10, f'dev_weekend_{doc}')
            model.AddAbsEquality(dev_weekend, weekend_count[doc] - 1)
            penalties.append(dev_weekend * 100)

        # ペナルティ3: OC2回から乖離
        for doc in self.g0:
            dev_oc = model.NewIntVar(0, 10, f'dev_oc_{doc}')
            model.AddAbsEquality(dev_oc, oc_count[doc] - 2)
            penalties.append(dev_oc * 50)

        # ペナルティ4: 土日祝2回かつ両方が日祝(土曜なし)
        for doc in self.doctors:
            sun_hol_count = sum(x[doc][s] for s in self.sunday_holiday_nichoku)
            sat_count = sum(x[doc][s] for s in self.saturday_nichoku)
            # 土日祝2回以上で土曜が0回の場合ペナルティ
            has_weekend_2plus = model.NewBoolVar(f'has_weekend_2plus_{doc}')
            model.Add(weekend_count[doc] >= 2).OnlyEnforceIf(has_weekend_2plus)
            model.Add(weekend_count[doc] < 2).OnlyEnforceIf(has_weekend_2plus.Not())

            no_saturday = model.NewBoolVar(f'no_saturday_{doc}')
            model.Add(sat_count == 0).OnlyEnforceIf(no_saturday)
            model.Add(sat_count >= 1).OnlyEnforceIf(no_saturday.Not())

            both_sun_hol = model.NewBoolVar(f'both_sun_hol_{doc}')
            model.AddBoolAnd([has_weekend_2plus, no_saturday]).OnlyEnforceIf(both_sun_hol)
            model.AddBoolOr([has_weekend_2plus.Not(), no_saturday.Not()]).OnlyEnforceIf(both_sun_hol.Not())

            penalties.append(both_sun_hol * 200)

        # ペナルティ5: 勤務間隔が短い場合
        sorted_shifts = sorted(self.shift_cols,
                               key=lambda s: (self.shift_day.get(s, 0),
                                             self.shift_sub.get(s, 0)))
        for doc in self.doctors:
            for i, s1 in enumerate(sorted_shifts):
                for s2 in sorted_shifts[i+1:]:
                    day1 = self.shift_day.get(s1, 0)
                    day2 = self.shift_day.get(s2, 0)
                    gap = day2 - day1
                    if gap < 4 and gap > 0:
                        # 両方に割り当てられたらペナルティ
                        both_assigned = model.NewBoolVar(f'gap_{doc}_{s1}_{s2}')
                        model.AddBoolAnd([x[doc][s1], x[doc][s2]]).OnlyEnforceIf(both_assigned)
                        model.AddBoolOr([x[doc][s1].Not(), x[doc][s2].Not()]).OnlyEnforceIf(both_assigned.Not())
                        penalty_value = (4 - gap) * 10
                        penalties.append(both_assigned * penalty_value)

        # 目的関数: ペナルティの合計を最小化
        model.Minimize(sum(penalties))

        # === 求解 ===
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = time_limit
        solver.parameters.num_search_workers = 4

        print(f"\n=== 最適化開始 (制限時間: {time_limit}秒) ===")
        status = solver.Solve(model)

        if status == cp_model.OPTIMAL:
            print("最適解が見つかりました!")
        elif status == cp_model.FEASIBLE:
            print("実行可能解が見つかりました (最適ではない可能性あり)")
        else:
            print("解が見つかりませんでした。制約を緩和してください。")
            return None

        # === 結果収集 ===
        result = {
            'duty': defaultdict(dict),
            'oncall': {},
            'duty_count': defaultdict(int),
            'oc_count': defaultdict(int),
            'weekday_count': defaultdict(int),
            'weekend_count': defaultdict(int),
        }

        for doc in self.doctors:
            for shift in self.shift_cols:
                if solver.Value(x[doc][shift]) == 1:
                    group = 'G0' if doc in self.g0 else 'G1'
                    result['duty'][shift][doc] = group
                    result['duty_count'][doc] += 1
                    if shift in self.weekday_shifts:
                        result['weekday_count'][doc] += 1
                    else:
                        result['weekend_count'][doc] += 1

        for doc in self.g0:
            for shift in self.shift_cols:
                if solver.Value(oc[doc][shift]) == 1:
                    result['oncall'][shift] = doc
                    result['oc_count'][doc] += 1

        return result

    def generate_output(self, result: dict, output_prefix: str = None):
        """結果をCSV/Excelに出力"""
        if output_prefix is None:
            output_prefix = f"{self.year}_{self.month}月_schedule"

        # Schedule DataFrame
        rows = []
        for shift in self.shift_cols:
            duty_g0 = ', '.join([d for d, g in result['duty'].get(shift, {}).items() if g == 'G0'])
            duty_g1 = ', '.join([d for d, g in result['duty'].get(shift, {}).items() if g == 'G1'])
            oc = result['oncall'].get(shift, '')
            rows.append({
                'Shift': shift,
                'Duty_G0': duty_g0,
                'Duty_G1': duty_g1,
                'Oncall_G0': oc
            })
        schedule_df = pd.DataFrame(rows)

        # Summary DataFrame
        summary_rows = []
        for doc in self.doctors:
            summary_rows.append({
                'Name': doc,
                'Group': self.group_map.get(doc, ''),
                'Duty_Weekday': result['weekday_count'].get(doc, 0),
                'Duty_Weekend': result['weekend_count'].get(doc, 0),
                'Duty_Total': result['duty_count'].get(doc, 0),
                'OC': result['oc_count'].get(doc, 0),
                'Total': result['duty_count'].get(doc, 0) + result['oc_count'].get(doc, 0)
            })
        summary_df = pd.DataFrame(summary_rows)

        # Excel出力
        excel_path = f"{output_prefix}.xlsx"
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            schedule_df.to_excel(writer, index=False, sheet_name='Schedule')
            summary_df.to_excel(writer, index=False, sheet_name='Summary')
        print(f"Excel出力: {excel_path}")

        # 元CSVに割り当て結果を書き込み
        df_annot = self.df_raw.copy()
        for shift, doctors in result['duty'].items():
            for doc in doctors:
                if doc in df_annot['Name'].values:
                    df_annot.loc[df_annot['Name'] == doc, shift] = 3  # Duty=3

        for shift, doc in result['oncall'].items():
            if doc in df_annot['Name'].values:
                df_annot.loc[df_annot['Name'] == doc, shift] = 4  # OC=4

        csv_path = f"{output_prefix}_annotated.csv"
        df_annot.to_csv(csv_path, index=False, encoding='cp932')
        print(f"CSV出力: {csv_path}")

        # カレンダー形式Excel
        self._generate_calendar(result, f"{output_prefix}_calendar.xlsx")

        # サマリー表示
        print("\n=== 割り当て結果サマリー ===")
        print(summary_df.to_string(index=False))

        return schedule_df, summary_df

    def _generate_calendar(self, result: dict, output_path: str):
        """カレンダー形式のExcel出力"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Calendar"

        # スタイル定義
        thin = Side(style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        hdr_fill = PatternFill("solid", fgColor="DDDDDD")
        hdr_font = Font(bold=True)
        red_font = Font(color="FF0000", bold=True)
        blue_font = Font(color="0000FF", bold=True)

        # タイトル
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
        title_cell = ws.cell(1, 1, f"{self.year}年{self.month}月 当番表")
        title_cell.alignment = center
        title_cell.font = Font(bold=True, size=16)

        # 曜日ヘッダ
        week = ["日", "月", "火", "水", "木", "金", "土"]
        for col_idx, wd in enumerate(week, 1):
            c = ws.cell(2, col_idx, wd)
            c.alignment = center
            c.font = hdr_font
            c.fill = hdr_fill
            c.border = border
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

        for r in range(3, 10):
            ws.row_dimensions[r].height = 60

        # 日付ごとの情報を集計
        day_info = defaultdict(lambda: {"日直": [], "当直": [], "OC": []})
        for shift, doctors in result['duty'].items():
            day = self.shift_day.get(shift)
            sub = self.shift_sub.get(shift)
            if day is None:
                continue

            names = "/".join(doctors.keys())
            if sub == 1:
                day_info[day]["日直"].append(names)
            else:
                day_info[day]["当直"].append(names)
                if shift in result['oncall']:
                    day_info[day]["OC"].append(result['oncall'][shift])

        # カレンダー本体
        first_date = datetime.date(self.year, self.month, 1)
        first_wd = first_date.weekday()
        col_offset = (first_wd + 1) % 7

        day = 1
        for row_idx in range(3, 9):
            for col_idx in range(1, 8):
                current_day = day - col_offset
                if current_day > 0:
                    try:
                        date = datetime.date(self.year, self.month, current_day)
                        cell = ws.cell(row_idx, col_idx)
                        cell.alignment = center
                        cell.border = border

                        info = day_info[current_day]
                        lines = []
                        if info["日直"]:
                            lines.append(f"日直: {', '.join(info['日直'])}")
                        if info["当直"]:
                            line = f"当直: {', '.join(info['当直'])}"
                            if info["OC"]:
                                line += f"\n(OC: {', '.join(info['OC'])})"
                            lines.append(line)

                        text = f"{current_day}\n" + "\n".join(lines)
                        cell.value = text

                        if date in self.holidays or date.weekday() == 6:
                            cell.font = red_font
                        elif date.weekday() == 5:
                            cell.font = blue_font
                    except ValueError:
                        pass
                day += 1

        wb.save(output_path)
        print(f"カレンダー出力: {output_path}")


def main():
    if len(sys.argv) < 4:
        print("使い方: python scheduler.py <CSV_FILE> <YEAR> <MONTH>")
        print("例: python scheduler.py 2026_1月_shift_schedule.csv 2026 1")
        sys.exit(1)

    csv_path = sys.argv[1]
    year = int(sys.argv[2])
    month = int(sys.argv[3])

    if not Path(csv_path).exists():
        print(f"エラー: ファイルが見つかりません: {csv_path}")
        sys.exit(1)

    scheduler = DutyScheduler(csv_path, year, month)
    result = scheduler.solve(time_limit=120)

    if result:
        scheduler.generate_output(result)
        print("\n完了しました!")


if __name__ == "__main__":
    main()
